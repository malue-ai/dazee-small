#!/usr/bin/env python3
"""
小搭子 vs clawdbot 对比测试 — 测试数据生成脚本

生成以下测试数据：
1. messy_sales.xlsx  — 格式混乱的销售数据 Excel（日期格式不统一、空行、中文列名）
2. mixed_files/      — 100 个混合类型文件（模拟文件整理场景）

使用方法：
    source /Users/liuyi/Documents/langchain/liuy/bin/activate
    cd docs/benchmark/data
    python generate_test_data.py
"""

import os
import json
import random
import datetime
from pathlib import Path


def generate_messy_excel():
    """生成格式混乱的销售数据 Excel

    特意制造以下问题：
    1. 日期格式不统一（2025-01-01 / 2025/1/1 / 1月1日 / Jan 1, 2025）
    2. 空行（第 5、12、20 行为空）
    3. 中文列名
    4. 数字和文本混合（"¥1,234" vs 1234）
    5. 缺失值（部分单元格为空）
    6. 类型不一致（同一列混合 int 和 str）
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("⚠️  需要安装 openpyxl: pip install openpyxl")
        print("   尝试生成 CSV 替代文件...")
        generate_messy_csv()
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "销售数据"

    # 标题行 — 合并单元格
    ws.merge_cells('A1:F1')
    ws['A1'] = '2024年度销售数据汇总表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # 空行
    # 第2行留空

    # 列标题（第3行）
    headers = ['日期', '产品名称', '销售额', '数量', '地区', '备注']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFE0E0E0', fill_type='solid')

    # 产品列表
    products = ['笔记本电脑', '无线鼠标', '机械键盘', 'USB-C 转接头',
                '显示器', '耳机', '移动硬盘', '摄像头']
    regions = ['华东', '华北', '华南', '西南', '华中']
    notes = ['正常', '促销', '退货', '换货', '', '预售', '团购', None]

    # 多种日期格式（故意制造混乱）
    date_formats = [
        lambda d: d.strftime('%Y-%m-%d'),       # 2024-01-15
        lambda d: d.strftime('%Y/%m/%d'),        # 2024/01/15
        lambda d: f"{d.month}月{d.day}日",       # 1月15日
        lambda d: d.strftime('%m/%d/%Y'),        # 01/15/2024
        lambda d: d.strftime('%Y年%m月%d日'),     # 2024年01月15日
        lambda d: str(d.day) + '/' + str(d.month),  # 15/1（歧义格式）
    ]

    # 生成数据行（从第4行开始）
    row = 4
    empty_rows = {8, 15, 23}  # 故意插入空行的位置
    start_date = datetime.date(2024, 1, 1)

    for i in range(30):
        if row in empty_rows:
            row += 1  # 跳过空行

        # 日期：随机选择格式
        date = start_date + datetime.timedelta(days=random.randint(0, 364))
        fmt = random.choice(date_formats)
        date_str = fmt(date)
        ws.cell(row=row, column=1, value=date_str)

        # 产品
        ws.cell(row=row, column=2, value=random.choice(products))

        # 销售额：混合格式（数字 vs 带符号字符串）
        amount = random.randint(500, 50000)
        if random.random() < 0.3:
            # 30% 概率使用文本格式 "¥1,234"
            ws.cell(row=row, column=3, value=f"¥{amount:,}")
        elif random.random() < 0.1:
            # 10% 概率留空
            pass
        else:
            ws.cell(row=row, column=3, value=amount)

        # 数量：混合类型
        qty = random.randint(1, 100)
        if random.random() < 0.2:
            ws.cell(row=row, column=4, value=str(qty) + "个")  # 文本格式
        else:
            ws.cell(row=row, column=4, value=qty)

        # 地区
        ws.cell(row=row, column=5, value=random.choice(regions))

        # 备注
        note = random.choice(notes)
        if note:
            ws.cell(row=row, column=6, value=note)

        row += 1

    # 合并单元格（制造额外混乱）
    ws.merge_cells('E10:E12')
    ws['E10'] = '华东地区'
    ws['E10'].alignment = Alignment(vertical='center')

    # 底部汇总行（文本格式的数字）
    row += 1
    ws.cell(row=row, column=1, value='合计')
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value='详见附表')  # 文本而非数字

    output_path = Path(__file__).parent / 'messy_sales.xlsx'
    wb.save(str(output_path))
    print(f"✅ 生成 messy_sales.xlsx ({output_path})")
    print(f"   - 30 条数据 + 3 个空行 + 合并单元格")
    print(f"   - 6 种日期格式混合")
    print(f"   - 销售额混合数字/文本格式")
    print(f"   - 数量列混合 int/str")


def generate_messy_csv():
    """openpyxl 不可用时的 CSV 替代方案"""
    import csv

    products = ['笔记本电脑', '无线鼠标', '机械键盘', 'USB-C 转接头',
                '显示器', '耳机', '移动硬盘', '摄像头']
    regions = ['华东', '华北', '华南', '西南', '华中']
    date_formats = [
        lambda d: d.strftime('%Y-%m-%d'),
        lambda d: d.strftime('%Y/%m/%d'),
        lambda d: f"{d.month}月{d.day}日",
        lambda d: d.strftime('%Y年%m月%d日'),
    ]

    output_path = Path(__file__).parent / 'messy_sales.csv'
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['2024年度销售数据汇总表', '', '', '', '', ''])
        writer.writerow([])  # 空行
        writer.writerow(['日期', '产品名称', '销售额', '数量', '地区', '备注'])

        start_date = datetime.date(2024, 1, 1)
        for i in range(30):
            if i in {4, 11, 19}:
                writer.writerow([])  # 空行
            date = start_date + datetime.timedelta(days=random.randint(0, 364))
            fmt = random.choice(date_formats)
            amount = random.randint(500, 50000)
            qty = random.randint(1, 100)
            amount_str = f"¥{amount:,}" if random.random() < 0.3 else str(amount)
            qty_str = f"{qty}个" if random.random() < 0.2 else str(qty)
            note = random.choice(['正常', '促销', '退货', '', '团购', ''])
            writer.writerow([fmt(date), random.choice(products), amount_str,
                             qty_str, random.choice(regions), note])

        writer.writerow(['合计', '', '详见附表', '', '', ''])

    print(f"✅ 生成 messy_sales.csv ({output_path})（openpyxl 不可用，使用 CSV 替代）")


def generate_mixed_files():
    """生成 100 个混合类型文件，用于文件整理和重命名测试

    文件类型分布：
    - .txt  (30个)：笔记、备忘
    - .md   (20个)：Markdown 文档
    - .csv  (15个)：小型数据表
    - .json (15个)：配置/数据文件
    - .log  (20个)：日志文件

    文件名故意不规范：
    - 混合中英文
    - 无序编号
    - 日期格式不统一
    - 包含空格和特殊字符
    """
    mixed_dir = Path(__file__).parent / 'mixed_files'
    mixed_dir.mkdir(exist_ok=True)

    # 清空已有文件
    for f in mixed_dir.iterdir():
        if f.is_file():
            f.unlink()

    created = 0
    categories = {
        'notes': [],
        'docs': [],
        'data': [],
        'config': [],
        'logs': [],
    }

    # ---- .txt 文件（30个）：笔记和备忘 ----
    note_topics = [
        '会议纪要', '周报', '待办事项', '读书笔记', '灵感记录',
        '项目计划', '问题记录', '学习笔记', '旅行计划', '购物清单',
    ]
    for i in range(30):
        topic = random.choice(note_topics)
        date = datetime.date(2024, random.randint(1, 12), random.randint(1, 28))

        # 故意使用不规范的文件名
        name_patterns = [
            f"{topic}_{date.strftime('%Y%m%d')}.txt",
            f"{date.strftime('%m月%d日')} {topic}.txt",
            f"note_{i+1}.txt",
            f"{topic}({i+1}).txt",
            f"{date.year}-{date.month}-{date.day} {topic}.txt",
        ]
        filename = random.choice(name_patterns)

        content = f"# {topic}\n\n日期：{date}\n\n这是一份{topic}的内容。\n"
        content += f"创建时间：{date.isoformat()}\n"
        content += f"分类：笔记\n"

        (mixed_dir / filename).write_text(content, encoding='utf-8')
        categories['notes'].append(filename)
        created += 1

    # ---- .md 文件（20个）：文档 ----
    doc_topics = [
        '技术方案', '设计文档', '需求分析', '测试报告', '用户手册',
        'API文档', '部署指南', '变更日志', '架构设计', '评审意见',
    ]
    for i in range(20):
        topic = random.choice(doc_topics)
        date = datetime.date(2024, random.randint(1, 12), random.randint(1, 28))

        name_patterns = [
            f"{topic}_v{random.randint(1,5)}.{random.randint(0,9)}.md",
            f"doc_{topic}_{date.strftime('%Y%m%d')}.md",
            f"{date.strftime('%Y-%m-%d')}_{topic}.md",
            f"【{topic}】最终版.md",
        ]
        filename = random.choice(name_patterns)

        content = f"# {topic}\n\n## 概述\n\n本文档描述了{topic}的相关内容。\n\n"
        content += f"## 日期\n\n{date.isoformat()}\n\n"
        content += f"## 内容\n\n详细内容待补充...\n"

        (mixed_dir / filename).write_text(content, encoding='utf-8')
        categories['docs'].append(filename)
        created += 1

    # ---- .csv 文件（15个）：数据表 ----
    data_types = ['用户数据', '销售报表', '库存清单', '考勤记录', '成绩单']
    for i in range(15):
        dtype = random.choice(data_types)
        date = datetime.date(2024, random.randint(1, 12), random.randint(1, 28))

        name_patterns = [
            f"{dtype}_{date.strftime('%Y%m%d')}.csv",
            f"data_{i+1}.csv",
            f"export_{dtype}.csv",
        ]
        filename = random.choice(name_patterns)

        # 生成简单 CSV 内容
        headers = ['ID', '名称', '数值', '日期']
        rows = []
        for j in range(random.randint(5, 20)):
            rows.append(f"{j+1},{dtype}_{j+1},{random.randint(100,9999)},{date.isoformat()}")

        content = ','.join(headers) + '\n' + '\n'.join(rows) + '\n'

        (mixed_dir / filename).write_text(content, encoding='utf-8')
        categories['data'].append(filename)
        created += 1

    # ---- .json 文件（15个）：配置 ----
    config_types = ['settings', 'config', 'preferences', 'theme', 'profile']
    for i in range(15):
        ctype = random.choice(config_types)

        name_patterns = [
            f"{ctype}_{i+1}.json",
            f"app_{ctype}.json",
            f"{ctype}_backup.json",
        ]
        filename = random.choice(name_patterns)

        data = {
            "type": ctype,
            "version": f"{random.randint(1,3)}.{random.randint(0,9)}",
            "created": datetime.date(2024, random.randint(1, 12),
                                     random.randint(1, 28)).isoformat(),
            "settings": {
                "key1": random.choice(["value_a", "value_b", "value_c"]),
                "key2": random.randint(1, 100),
                "enabled": random.choice([True, False]),
            }
        }

        (mixed_dir / filename).write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
        categories['config'].append(filename)
        created += 1

    # ---- .log 文件（20个）：日志 ----
    log_levels = ['INFO', 'WARNING', 'ERROR', 'DEBUG']
    for i in range(20):
        date = datetime.date(2024, random.randint(1, 12), random.randint(1, 28))

        name_patterns = [
            f"app_{date.strftime('%Y%m%d')}.log",
            f"server_{i+1}.log",
            f"error_{date.strftime('%m%d')}.log",
            f"debug.log.{i+1}",
        ]
        filename = random.choice(name_patterns)

        lines = []
        for j in range(random.randint(10, 50)):
            ts = datetime.datetime(2024, date.month, date.day,
                                   random.randint(0, 23),
                                   random.randint(0, 59),
                                   random.randint(0, 59))
            level = random.choice(log_levels)
            msg = random.choice([
                "请求处理完成",
                "数据库连接成功",
                "缓存未命中",
                "文件读取失败",
                "用户登录",
                "任务调度执行",
                "超时重试",
            ])
            lines.append(f"[{ts.isoformat()}] [{level}] {msg}")

        content = '\n'.join(lines) + '\n'

        (mixed_dir / filename).write_text(content, encoding='utf-8')
        categories['logs'].append(filename)
        created += 1

    # 输出统计
    print(f"✅ 生成 {created} 个混合文件到 mixed_files/")
    for cat, files in categories.items():
        print(f"   - {cat}: {len(files)} 个文件")

    # 生成文件清单（用于测试验证）
    manifest_path = mixed_dir / '_manifest.json'
    manifest_path.write_text(
        json.dumps(categories, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"   - 清单文件: {manifest_path}")


def main():
    """一键生成所有测试数据"""
    print("=" * 60)
    print("小搭子 vs clawdbot 对比测试 — 测试数据生成")
    print("=" * 60)
    print()

    # 设置随机种子（确保可重复生成）
    random.seed(42)

    print("[1/2] 生成格式混乱 Excel...")
    generate_messy_excel()
    print()

    print("[2/2] 生成混合文件集...")
    generate_mixed_files()
    print()

    print("=" * 60)
    print("全部测试数据生成完成！")
    print()
    print("手动准备项：")
    print("  - 扫描 PDF：参见 scanned_pdf_note.md")
    print("=" * 60)


if __name__ == '__main__':
    main()
